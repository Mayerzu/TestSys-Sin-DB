from openpyxl import load_workbook
from openpyxl.utils import column_index_from_string
from services.excel.excel_store import set_excel_data
from utils.login_validation import is_valid_email
from utils.files.file_config import EXCEL_TEST_PATH, EXCEL_TEST_COLUMNS
from utils.replace_accents import replace_accents
from utils.excel_validations import *
from datetime import datetime
import os

def get_excel_data():
   excel_path = EXCEL_TEST_PATH

   if not os.path.exists(excel_path): return {
      'success': False,
      'message': f'Archivo ({excel_path}) no encontrado'
   }

   wb = load_workbook(filename=excel_path, data_only=True)
   sheet = wb.active

   if sheet.max_column < EXCEL_TEST_COLUMNS:
      wb.close()
      os.remove(EXCEL_TEST_PATH)

      return {
         'success': False,
         'message': f'El archivo excel no tiene las columnas necesarias ({EXCEL_TEST_COLUMNS} columnas)'
      }

   init_col_user = column_index_from_string('A')
   end_col_user = column_index_from_string('H')
   init_col_score = column_index_from_string('DY')
   end_col_score = column_index_from_string('EH')
   
   areas = [
      'Servicio Social', 'Persuasivo', 'Literario', 'Artístico Plástico', 'Musical',
      'Trabajo de Oficina', 'Científico', 'Cálculo', 'Mecánico', 'Aire Libre'
   ]

   data = []

   for row in range(2, sheet.max_row+1):
      values = [sheet.cell(row=row, column=col).value for col in range(init_col_user, end_col_user+1)]

      if values[0]:
         scores = [sheet.cell(row=row, column=col).value for col in range(init_col_score, end_col_score+1)]
         area_scores = {}

         for index, area in enumerate(areas):
            fixed_area = '_'.join(replace_accents(area.strip().lower()).split(' '))
            
            area_scores[fixed_area] = { 
               'area': fixed_area,
               'area_name': area.strip(),
               'score': scores[index]
            }

         evaluation_date = values[0].strftime('%d/%m/%Y').strip()
         default_email = str(values[1]).strip() if values[1] else ''
         name = str(values[2]).strip().title() if values[2] else ''

         excel_email = str(values[3]).strip() if values[3] else ''
         excel_phone = str(values[4]).strip() if values[4] else ''
         excel_age = values[5] if values[5] else ''
         excel_birth_date = values[6] if values[6] else ''
         careers = values[7] if values[7] else ''

         age = int(extract_digits(excel_age))
         phone = valid_phone(excel_phone)
         email = excel_email if is_valid_email(excel_email) else default_email
         birth_date = valid_birth_date(excel_birth_date)

         index = len(data) + 1

         user = {
            'index': index,
            'evaluation_date': evaluation_date,
            'default_email': default_email,
            'name': name,
            'user_email': excel_email,
            'phone': phone,
            'age': age,
            'birth_date': birth_date,
            'careers': careers,
            'scores': area_scores,
            'email': email
         }

         data.append(user)
   
   wb.close()

   sorted_data = sorted(data, key=lambda x: datetime.strptime(x['evaluation_date'], '%d/%m/%Y'), reverse=True)
   data = []

   for index, user in enumerate(sorted_data, start=1):
      user['index'] = index
      data.append(user)

   set_excel_data(data)

   return data